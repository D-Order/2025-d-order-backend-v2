# -*- coding: utf-8 -*-
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_FMT = "yyyy-mm-dd hh:mm"
HEADERS = ["코드", "발급테이블", "사용여부", "사용시각"]  # 생성일 제거, 테이블 추가

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 40)

def _table_label(table_obj) -> str:
    """테이블 표시값: name/number가 있으면 그걸 쓰고, 없으면 id"""
    if table_obj is None:
        return ""
    # 프로젝트 Table 모델에 맞춰 우선순위로 표시 (필요하면 필드명 바꿔도 됨)
    for attr in ("name", "table_name", "number", "table_no", "id"):
        if hasattr(table_obj, attr) and getattr(table_obj, attr) is not None:
            return str(getattr(table_obj, attr))
    return ""

def build_codes_only_xlsx(qs, sheet_name: str = "codes", meta_title: Optional[str] = None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 데이터 (CouponCode 모델: code, issued_to_table, used_at)
    # 사용여부는 used_at 존재 여부로 판단
    for code in qs:
        table_label = _table_label(code.issued_to_table)
        ws.append([
            code.code,
            table_label,
            "Y" if code.used_at else "N",
            code.used_at,
        ])

    # 날짜 포맷 (사용시각 = 4번째 컬럼)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=4)
        if cell.value:
            cell.number_format = DATE_FMT

    _autosize(ws)

    # 메타 시트(선택)
    if meta_title:
        meta = wb.create_sheet("meta")
        meta.append(["Title", meta_title])
        _autosize(meta)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
